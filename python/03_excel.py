# python/03_excel.py

import pandas as pd
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

# Directories
root_dir = os.path.dirname(os.path.dirname(__file__))

output_dir = os.path.join(root_dir, "outputs")
report_dir = os.path.join(root_dir, "reports")

os.makedirs(report_dir, exist_ok=True)

# Loading exported KPI files
summary_df = pd.read_csv(os.path.join(output_dir, "kpi_summary.csv"))
monthly_df = pd.read_csv(os.path.join(output_dir, "monthly_accuracy.csv"))
weekly_df = pd.read_csv(os.path.join(output_dir, "weekly_kpi.csv"))
sku_df = pd.read_csv(os.path.join(output_dir, "sku_error_analysis.csv"))
warehouse_df = pd.read_csv(os.path.join(output_dir, "warehouse_defects_analysis.csv"))
supplier_df = pd.read_csv(os.path.join(output_dir, "supplier_performance.csv"))
region_df = pd.read_csv(os.path.join(output_dir, "regional_analysis.csv"))
severity_df = pd.read_csv(os.path.join(output_dir, "variance_severity.csv"))
high_risk_df = pd.read_csv(os.path.join(output_dir, "kpi_high_risk_skus.csv"))
problematic_df = pd.read_csv(os.path.join(output_dir, "top_problematic_SKU.csv"))
promo_df = pd.read_csv(os.path.join(output_dir, "promotion_impact.csv"))
lead_time_df = pd.read_csv(os.path.join(output_dir, "lead_time_defects.csv"))
corrective_df = pd.read_csv(os.path.join(output_dir, "corrective_actions.csv"))

# Workbook
wb = Workbook()

# Colors
DARK_BLUE = "111184"
LIGHT_BLUE = "90D5FF"
RED = "880808"
YELLOW = "FFC000"
GREEN = "008000"

# Styling functions
def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def style_good(cell):
    cell.fill = PatternFill("solid", fgColor=GREEN)

def style_warning(cell):
    cell.fill = PatternFill("solid", fgColor=YELLOW)

def style_critical(cell):
    cell.fill = PatternFill("solid", fgColor=RED)

def thin_border():
    side = Side(style="thin", color="D9D9D9")

    return Border(
        left=side,
        right=side,
        top=side,
        bottom=side
    )

def write_dataframe(ws, df, start_row=1, start_col=1):

    # headers
    for c, column in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=c)
        cell.value = column
        style_header(cell)
        cell.border = thin_border()

    # rows
    for r, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c, value in enumerate(row, start=start_col):
            cell = ws.cell(row=r, column=c)
            cell.value = value
            cell.border = thin_border()

def auto_fit_columns(ws):

    for column_cells in ws.columns:

        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 4


# executive summary page
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.merge_cells("A1:F1")
ws1["A1"] = "ICQA Inventory Accuracy Dashboard"
ws1["A1"].font = Font(bold=True, size=16)
ws1["A1"].alignment = Alignment(horizontal="center")

write_dataframe(ws1, summary_df, start_row=3)

# Highlighting KPI warnings
for row in range(4, len(summary_df) + 4):
    kpi_name = ws1.cell(row=row, column=1).value
    value = ws1.cell(row=row, column=2).value
    if "%" in kpi_name:
        if value >= 95:
            style_good(ws1.cell(row=row, column=2))
        elif value < 93:
            style_warning(ws1.cell(row=row, column=2))

# monthly accuracy trend page
ws2 = wb.create_sheet("Monthly Trend")
write_dataframe(ws2, monthly_df)

# Accuracy chart
line_chart = LineChart()
line_chart.title = "Monthly Inventory Accuracy"
line_chart.height = 10
line_chart.width = 18

data = Reference(
    ws2,
    min_col=3,
    min_row=1,
    max_row=len(monthly_df) + 1
)

line_chart.add_data(data, titles_from_data=True)

categories = Reference(
    ws2,
    min_col=1,
    min_row=2,
    max_row=len(monthly_df) + 1
)

line_chart.set_categories(categories)

ws2.add_chart(line_chart, "J2")

# weekly kpi page
ws3 = wb.create_sheet("Weekly KPI")
write_dataframe(ws3, weekly_df)

# SKU analysis page
ws4 = wb.create_sheet("SKU Error Analysis")
write_dataframe(ws4, sku_df)

for row in range(2, len(sku_df) + 2):
    defect_rate = ws4.cell(row=row, column=4).value
    if defect_rate >= 10:
        for col in range(1, len(sku_df.columns) + 1):
            style_critical(ws4.cell(row=row, column=col))
    elif defect_rate >= 5:
        for col in range(1, len(sku_df.columns) + 1):
            style_warning(ws4.cell(row=row, column=col))

# Top defect chart
bar_chart = BarChart()

bar_chart.title = "Top SKU Defect Rate"
bar_chart.height = 12
bar_chart.width = 20

data = Reference(
    ws4,
    min_col=4,
    min_row=1,
    max_row=min(len(sku_df) + 1, 15)
)

categories = Reference(
    ws4,
    min_col=1,
    min_row=2,
    max_row=min(len(sku_df) + 1, 15)
)

bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)

ws4.add_chart(bar_chart, "M2")

# Warehouse analysis page
ws5 = wb.create_sheet("Warehouse Analysis")
write_dataframe(ws5, warehouse_df)

for row in range(2, len(warehouse_df) + 2):
    defect_rate = ws5.cell(row=row, column=4).value
    if defect_rate >= 3.5:
        for col in range(1, len(warehouse_df.columns) + 1):
            style_warning(ws5.cell(row=row, column=col))

# supplier performance page
ws6 = wb.create_sheet("Supplier Performance")
write_dataframe(ws6, supplier_df)

for row in range(2, len(supplier_df) + 2):
    lead_time = ws6.cell(row=row, column=2).value
    if lead_time >= 5:
        for col in range(1, len(supplier_df.columns) + 1):
            style_warning(ws6.cell(row=row, column=col))

# regional analysis page
ws7 = wb.create_sheet("Regional Analysis")
write_dataframe(ws7, region_df)

# variance severity page
ws8 = wb.create_sheet("Variance Severity")
write_dataframe(ws8, severity_df)

for row in range(2, len(severity_df) + 2):
    severity = ws8.cell(row=row, column=1).value
    if severity == "Critical":
        for col in range(1, len(severity_df.columns) + 1):
            style_critical(ws8.cell(row=row, column=col))

    elif severity == "Moderate":
        for col in range(1, len(severity_df.columns) + 1):
            style_warning(ws8.cell(row=row, column=col))

# hgh risk SKUs page
ws9 = wb.create_sheet("High Risk SKUs")
write_dataframe(ws9, high_risk_df)

for row in range(2, len(high_risk_df) + 2):
    defects = ws9.cell(row=row, column=4).value
    if defects >= 20:
        for col in range(1, len(high_risk_df.columns) + 1):
            style_critical(ws9.cell(row=row, column=col))

# Top problematic sku page
ws10 = wb.create_sheet("Top Problematic SKUs")
write_dataframe(ws10, problematic_df)

# Promotional impact page
ws11 = wb.create_sheet("Promotion Impact")
write_dataframe(ws11, promo_df)

# Lead time defects page
ws12 = wb.create_sheet("Lead Time Defects")
write_dataframe(ws12, lead_time_df)

# Corrective actions page
ws13 = wb.create_sheet("Corrective Actions")

write_dataframe(ws13, corrective_df)

for row in range(2, len(corrective_df) + 2):

    priority = ws13.cell(row=row, column=4).value

    if priority == "High":
        for col in range(1, len(corrective_df.columns) + 1):
            style_critical(ws13.cell(row=row, column=col))

    elif priority == "Medium":
        for col in range(1, len(corrective_df.columns) + 1):
            style_warning(ws13.cell(row=row, column=col))


for ws in wb.worksheets:
    auto_fit_columns(ws)

# Save workbook
report_path = os.path.join(report_dir,"ICQA_Inventory_Report.xlsx")
wb.save(report_path)
print(f"Excel report saved to: {report_path}")